import tempfile
import unittest
from pathlib import Path

import openpyxl

from sports_research.export.xlsx_export import export_xlsx
from sports_research.models.event_result import make_event_result
from sports_research.models.source import make_source
from sports_research.research.engine import ResearchOutcome


class TestExportXLSX(unittest.TestCase):
    def test_workbook_has_all_four_required_sheets(self):
        event = make_event_result(
            sport="football", competition="PL", season="2020-2021", date="2020-08-15",
            participants=[{"name": "A", "role": "home", "score": 1}, {"name": "B", "role": "away", "score": 0}],
            status="completed", result="home_win",
            source="s", source_url="https://example.invalid", source_accessed_at="2026-08-23T12:00:00Z",
        )
        source = make_source(source_id="src_1", title="t", url="https://example.invalid",
                              retrieved_at="2026-08-23T12:00:00Z", retrieval_status="ok", source_type="other")
        outcome = ResearchOutcome(
            request={"raw_query": "q", "status": "ready"}, plan={"search_queries": []},
            search_execution={"results": []}, records=[event], sources=[source],
            validation_problems={}, duplicate_groups=[],
            completeness={"expected": None, "found": 1, "duplicate": 0, "unresolved": 0, "missing": None},
        )

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            export_xlsx(outcome, path)
            wb = openpyxl.load_workbook(path)

        self.assertEqual(wb.sheetnames, ["Results", "Sources", "Research Summary", "Validation Issues"])
        self.assertEqual(wb["Results"].cell(row=2, column=1).value, event["event_id"])
        self.assertEqual(wb["Sources"].cell(row=2, column=3).value, "https://example.invalid")


if __name__ == "__main__":
    unittest.main()
