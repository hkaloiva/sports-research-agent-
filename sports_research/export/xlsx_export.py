"""Excel (.xlsx) export via openpyxl.

Sheet 1: Results (one row per event/participant, same shape as the CSV
export). Sheet 2: Sources. Sheet 3: Research Summary. Sheet 4: Validation
Issues.
"""

from openpyxl import Workbook

from .csv_export import _rows_for_event


def export_xlsx(outcome, path) -> None:
    wb = Workbook()

    results_ws = wb.active
    results_ws.title = "Results"
    result_rows = [row for event in outcome.records for row in _rows_for_event(event)]
    if result_rows:
        headers = list(result_rows[0].keys())
        results_ws.append(headers)
        for row in result_rows:
            results_ws.append([row.get(h) for h in headers])
    else:
        results_ws.append(["(no results extracted)"])

    sources_ws = wb.create_sheet("Sources")
    source_headers = ["source_id", "title", "url", "domain", "retrieved_at", "retrieval_status", "source_type"]
    sources_ws.append(source_headers)
    for source in outcome.sources:
        sources_ws.append([source.get(h) for h in source_headers])

    summary_ws = wb.create_sheet("Research Summary")
    summary_ws.append(["Field", "Value"])
    summary_ws.append(["Raw query", outcome.request.get("raw_query") if outcome.request else None])
    summary_ws.append(["Status", outcome.request.get("status") if outcome.request else None])
    summary_ws.append(["Queries executed", len(outcome.plan["search_queries"]) if outcome.plan else 0])
    summary_ws.append(["Sources found", len(outcome.search_execution["results"]) if outcome.search_execution else 0])
    summary_ws.append(["Sources retrieved", len(outcome.sources)])
    summary_ws.append(["Records extracted", len(outcome.records)])
    if outcome.completeness:
        for key, value in outcome.completeness.items():
            summary_ws.append([f"Completeness: {key}", value])
    summary_ws.append(["Duplicate groups found", len(outcome.duplicate_groups)])
    verified = sum(1 for r in outcome.records if r.get("verification_status") == "verified")
    conflicting = sum(1 for r in outcome.records if r.get("verification_status") == "conflicting")
    unverified = sum(1 for r in outcome.records if r.get("verification_status") == "unverified")
    summary_ws.append(["Verified records", verified])
    summary_ws.append(["Conflicting records", conflicting])
    summary_ws.append(["Unverified records", unverified])

    issues_ws = wb.create_sheet("Validation Issues")
    issues_ws.append(["event_id", "problem"])
    for event_id, problems in outcome.validation_problems.items():
        for problem in problems:
            issues_ws.append([event_id, problem])

    wb.save(path)
